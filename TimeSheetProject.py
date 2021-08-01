# -*- coding: utf-8 -*-
"""
Created on Fri Jul 31 16:59:30 2020

@author: Ori Simcha
"""

from selenium import webdriver
import openpyxl
from pathlib import Path
from tkinter import messagebox
import datetime
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.chrome.service import Service
from subprocess import CREATE_NO_WINDOW
import os.path

import json
from selenium.webdriver.common.by import By
import getpass
import time
import calendar
import pickle


class TimeSheet:

    def __init__(self, excel_path ):

        #self.folder_path = StringVar()
        # self.validate = Validation.Validation()
        self.excel_path =excel_path
        self.url = 'http://10.10.1.131/TimeSheet/login.aspx'       #http://193.17.74.188:90/TimeSheet'

    def execute_driver (self):

        options = webdriver.ChromeOptions()
        # options.add_argument("--disable-popup-blocking")
        options.add_argument("--headless")
        # prefs = {"download.default_directory": r"C:\New_Download"}
        # options.add_experimental_option("prefs", prefs)
        # options.add_argument("user-data-dir=C:\\Users\\"+getpass.getuser()+"\\AppData\\Local\\Google\\Chrome\\User Data\\Default")  # this is the directory for the cookies
        # options.add_argument('user-data-dir=C:/Users/oris/AppData/Local/Google/Chrome/User Data/Default')
        # options.add_argument("profile-directory=Default")
        # print('C:/Users/'+getpass.getuser()+'/chromedriver')


        service = Service('C:/Users/'+getpass.getuser()+'/chromedriver')
        service.creationflags = CREATE_NO_WINDOW


        if  os.path.isfile('C:/Users/'+getpass.getuser()+'/chromedriver.exe')== False:
            messagebox.showerror(title='TimeSheet Robot',  message='chromedriver.exe not exists in '+'C:/Users/'+getpass.getuser())
            # self.driver.quit()
            return 1


        self.driver = webdriver.Chrome('C:/Users/'+getpass.getuser()+'/chromedriver', options=options ,  service = service)#'C:/Users/oris/PycharmProjects/pythonProject/chromedriver', options=options)

        # self.driver.service.stop()

        self.driver.set_window_size(1500, 1500)
        self.driver.execute_script("document.body.style.zoom = 'zoom90%'")

        if  os.path.isfile('C:/Users/'+getpass.getuser()+'/cookies.json')== False :
            messagebox.showerror(title='TimeSheet Robot',  message='cookies.json file not exists in '+'C:/Users/'+getpass.getuser())
            # self.driver.quit()
            return 1

        #load_cookie(driver, url)
        cookies = json.load(open('C:/Users/'+getpass.getuser()+'/cookies.json',"r"))

        try:
         self.driver.get(self.url)
        except:
            messagebox.showerror(title='TimeSheet Robot', message='The system failed to connect to Time Sheet domain, Please check if you are connected to vpn')
            # self.driver.quit()
            return 1

        for cookie in cookies:
             if 'expiry' in cookie:
              cookie['expiry'] = int(cookie['expiry'])
             self.driver.add_cookie(cookie)

        self.driver.refresh()
        # time.sleep(1)


        idx = str(self.excel_path).rfind('/')

        excel_name = str(self.excel_path)[-(len(self.excel_path) - idx - 1):]
        path_name= str(self.excel_path)[:-(len(self.excel_path) - idx - 1)]



        xlsx_file = Path(path_name, excel_name)


        self.wb_obj = openpyxl.load_workbook(xlsx_file)
        # Read the active sheet:
        self.sheet = self.wb_obj.active



    def play(self):

        self.driver.find_element_by_id("ContentPlaceHolder1_btnAdd").click()

        time.sleep(2)


        for i , row in enumerate(self.sheet.iter_rows(values_only=True)):

            if i == 0:
                continue

            else:

             # for cell in row:

                if row[0] == None:
                        continue


                self.driver.find_element_by_id("ContentPlaceHolder1_AddReport1_DateTextBox").click()
                time.sleep(2)
                # self.driver.find_element_by_id("ContentPlaceHolder1_AddReport1_DateTextBox").clear()
                # time.sleep(2)

                
                dateformat = "%Y-%m-%d" #"%Y/%m/d"
                try:
                    validtime = datetime.datetime.strptime(str(row[2].date()), dateformat)

                except:
                    messagebox.showerror(title='TimeSheet Robot', message='the date in row ' + str(i+1) + ' are not in a valid format.')
                    # self.driver.quit()
                    return

                day = str(row[2].day)
                month_name = calendar.month_name[int(row[2].strftime("%m"))]

                try:
                    self.driver.find_element_by_xpath("//div[@id='ContentPlaceHolder1_AddReport1_up1']//table[@id='ContentPlaceHolder1_AddReport1_Calendar1']//a[@title="+"'"+month_name+" "+day+"']").click()
                except:
                    # self.driver.quit()
                    messagebox.showerror(title='TimeSheet Robot',message='the month is not the current month')


                # insert '03:25'
                time.sleep(2)
                #h , m = row[3].split(':')
                # print(row[3].minute[0])


                timeformat = "%H:%M:%S"
                try:
                  validtime = datetime.datetime.strptime(str(row[3]), timeformat)

                except :
                        messagebox.showerror(title='TimeSheet Robot', message='the time in row '+str(i+1)+' is not a valid format.')
                        # self.driver.quit()
                        return

                if str(row[3].minute) != '0':
                    #driver.find_element_by_xpath("//div//input[@class='MaskedEditError']").send_keys('25')
                    self.driver.find_element_by_xpath("//input[@name='ctl00$ContentPlaceHolder1$AddReport1$txtHours']").click()
                    self.driver.find_element_by_xpath("//input[@name='ctl00$ContentPlaceHolder1$AddReport1$txtHours']").send_keys(Keys.ARROW_RIGHT)
                    self.driver.find_element_by_xpath("//input[@name='ctl00$ContentPlaceHolder1$AddReport1$txtHours']").send_keys(str(row[3].minute)[1]) #'5'
                    self.driver.find_element_by_xpath("//input[@name='ctl00$ContentPlaceHolder1$AddReport1$txtHours']").send_keys(Keys.ARROW_RIGHT)
                    self.driver.find_element_by_xpath("//input[@name='ctl00$ContentPlaceHolder1$AddReport1$txtHours']").send_keys(Keys.ARROW_RIGHT)
                    self.driver.find_element_by_xpath("//input[@name='ctl00$ContentPlaceHolder1$AddReport1$txtHours']").send_keys(str(row[3].minute)[0]) #'2'
                    time.sleep(1)

                else:
                    # driver.find_element_by_xpath("//div//input[@class='MaskedEditError']").send_keys('25')
                    self.driver.find_element_by_xpath("//input[@name='ctl00$ContentPlaceHolder1$AddReport1$txtHours']").click()
                    self.driver.find_element_by_xpath("//input[@name='ctl00$ContentPlaceHolder1$AddReport1$txtHours']").send_keys(Keys.ARROW_RIGHT)
                    self.driver.find_element_by_xpath("//input[@name='ctl00$ContentPlaceHolder1$AddReport1$txtHours']").send_keys(0)  # '5'
                    self.driver.find_element_by_xpath("//input[@name='ctl00$ContentPlaceHolder1$AddReport1$txtHours']").send_keys(Keys.ARROW_RIGHT)
                    self.driver.find_element_by_xpath("//input[@name='ctl00$ContentPlaceHolder1$AddReport1$txtHours']").send_keys(Keys.ARROW_RIGHT)
                    self.driver.find_element_by_xpath("//input[@name='ctl00$ContentPlaceHolder1$AddReport1$txtHours']").send_keys(0)  # '2'
                    time.sleep(1)

                #driver.find_element_by_id("ctl00_ContentPlaceHolder1_AddReport1_txtHours").send_keys('0025')
                self.driver.find_element_by_xpath("//input[@name='ctl00$ContentPlaceHolder1$AddReport1$txtHours']").send_keys(Keys.ARROW_RIGHT)
                self.driver.find_element_by_xpath("//input[@name='ctl00$ContentPlaceHolder1$AddReport1$txtHours']").send_keys(Keys.ARROW_RIGHT)
                self.driver.find_element_by_xpath("//input[@name='ctl00$ContentPlaceHolder1$AddReport1$txtHours']").send_keys(Keys.ARROW_RIGHT)
                self.driver.find_element_by_xpath("//input[@name='ctl00$ContentPlaceHolder1$AddReport1$txtHours']").send_keys(Keys.ARROW_RIGHT)
                time.sleep(1)


                if(len(str(row[3].hour)) > 1):

                 print(str(row[3].hour))
                 self.driver.find_element_by_xpath("//input[@name='ctl00$ContentPlaceHolder1$AddReport1$txtHours']").send_keys(str(row[3].hour)[0]) #'0'
                 self.driver.find_element_by_xpath("//input[@name='ctl00$ContentPlaceHolder1$AddReport1$txtHours']").send_keys(str(row[3].hour)[1])  # '3'
                else:
                 self.driver.find_element_by_xpath("//input[@name='ctl00$ContentPlaceHolder1$AddReport1$txtHours']").send_keys(0)  # '0'
                 self.driver.find_element_by_xpath("//input[@name='ctl00$ContentPlaceHolder1$AddReport1$txtHours']").send_keys(str(row[3].hour)[0])  # '3'


                #driver.find_element_by_xpath("//div[@id='ctl00_ContentPlaceHolder1_AddReport1_Panel1']//input[@name='ctl00$ContentPlaceHolder1$AddReport1$txtHours']").send_keys(Keys.TAB)

                time.sleep(2)
                self.driver.find_element_by_id("ContentPlaceHolder1_AddReport1_FTC").click()
                time.sleep(2)
                self.driver.find_element_by_id("ContentPlaceHolder1_AddReport1_txtClient").send_keys(row[0])
                time.sleep(2)
                self.driver.find_element_by_xpath("//textarea[@name='ctl00$ContentPlaceHolder1$AddReport1$txtDesc']").click()
                time.sleep(2)

                if len(row[4]) >150:
                    messagebox.showerror(title='TimeSheet Robot', message='The Notes are more than 150 chars, pls change it')
                    # self.driver.quit()
                    return

                try:
                    self.driver.find_element_by_xpath( "//textarea[@name='ctl00$ContentPlaceHolder1$AddReport1$txtDesc']").send_keys(row[4])  # .decode("utf8")
                    # time.sleep(2)
                except:
                  messagebox.showerror(title='TimeSheet Robot', message='in row '+str(i+1)+' the Notes are empty')
                  # self.driver.quit()
                  return

                time.sleep(3)
                try:
                 self.driver.find_element_by_xpath("//input[@name='ctl00$ContentPlaceHolder1$AddReport1$txtProject']").send_keys(row[1])
                except:
                    messagebox.showerror(title='TimeSheet Robot', message='in row '+str(i+1)+' the SubProject Code is not exists')
                    # self.driver.quit()
                    return

                time.sleep(3)
                try:
                 self.driver.find_element_by_xpath("//input[@name='ctl00$ContentPlaceHolder1$AddReport1$txtHours']").send_keys(Keys.TAB)

                except:
                    messagebox.showerror(title='TimeSheet Robot', message='row: ' + str(i+1) + '  something go wrong, please try again')
                    # self.driver.quit()
                    return


                from selenium.webdriver.common.action_chains import ActionChains


                try:
                    self.driver.find_element_by_id("ContentPlaceHolder1_AddReport1_Button1").click()
                except:
                    messagebox.showerror(title='TimeSheet Robot', message='in row '+str(i+1)+' some of the values are invalid or less , pls check it')
                    # self.driver.quit()
                    return
                # driver.implicitly_wait(10)
                # ActionChains(driver).move_to_element(button).click(button).perform()
                time.sleep(3)
                self.driver.find_element_by_id("ContentPlaceHolder1_AddReport1_btnCancel").click()
                time.sleep(2)



        self.driver.quit()
        messagebox.showinfo(title='TimeSheet Robot', message='Upload completed successfully!')


    def clean_empty_rows(self):

        print(self.sheet.max_row)


        # iterate the sheet object
        for row in self.sheet:
            for cell in row:
                if cell.value != None:
                    break

                else:
                    print(row[0].row)
                    # get the row number from the first cell
                    # and remove the row
                    self.sheet.delete_rows(row[0].row ,row[0].row )
                    break

        print(self.sheet.max_row)

        try:
            self.wb_obj.save(self.excel_path)
        except:
            messagebox.showerror(title='TimeSheet Robot', message='The Excel file is open , pls close him first')
            return 1

        return 0


    def clean_excel(self, toclean):

        print(toclean)
        if(toclean == 1):

            # # load excel file
            # idx = str(self.excel_path).rfind('/')
            #
            # excel_name = str(self.excel_path)[-(len(self.excel_path) - idx - 1):]
            # path_name = str(self.excel_path)[:-(len(self.excel_path) - idx - 1)]
            #
            # xlsx_file = Path(path_name, excel_name)
            # # print(xlsx_file)
            # wb_obj = openpyxl.load_workbook(xlsx_file)
            #
            # # Read the active sheet:
            # # self.sheet = wb_obj.active
            # self.sheet = wb_obj.get_sheet_by_name("גיליון1")
            # print(self.sheet)
            # # for i, row in enumerate(self.sheet.iter_rows(values_only=True)):
            # # while (self.sheet.max_row > 1):
            #     # this method removes the row 2
            self.sheet.delete_rows(2 ,self.sheet.max_row+1)
            self.wb_obj.save(self.excel_path)

            # return to main function




